"""Excel (.xlsx/.xls) to Markdown converter for doc2md."""

import sys
from datetime import datetime
from pathlib import Path

from utils.tables import _format_cell_value, _sheets_to_markdown


def convert_excel_native(input_path, output_dir, config, args):
    """直接讀取 .xlsx，每個 Sheet 轉為 Markdown 表格。"""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl is required. Install with: pip install openpyxl", file=sys.stderr)
        return False

    try:
        wb = openpyxl.load_workbook(str(input_path), data_only=True)
    except Exception as e:
        print(f"Error opening {input_path}: {e}", file=sys.stderr)
        return False

    sheet_names = wb.sheetnames
    print(f"Processing: {input_path} ({len(sheet_names)} sheets)", file=sys.stderr)

    sheets_data = []
    for sheet_name in sheet_names:
        ws = wb[sheet_name]

        # Handle merged cells
        merged_values = {}
        for merged_range in ws.merged_cells.ranges:
            top_left_value = ws.cell(merged_range.min_row, merged_range.min_col).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    if row == merged_range.min_row and col == merged_range.min_col:
                        merged_values[(row, col)] = top_left_value
                    else:
                        merged_values[(row, col)] = ''

        all_rows = []
        for row in ws.iter_rows():
            cells = []
            for cell in row:
                key = (cell.row, cell.column)
                val = merged_values[key] if key in merged_values else cell.value
                cells.append(_format_cell_value(val))
            all_rows.append(cells)

        sheets_data.append((sheet_name, all_rows))

    wb.close()
    return _sheets_to_markdown(input_path, output_dir, config, args, sheets_data)


def convert_xls_native(input_path, output_dir, config, args):
    """直接讀取 .xls (Excel 97-2003)，每個 Sheet 轉為 Markdown 表格。"""
    try:
        import xlrd
    except ImportError:
        print("Error: xlrd is required. Install with: pip install xlrd", file=sys.stderr)
        return False

    try:
        wb = xlrd.open_workbook(str(input_path), formatting_info=True)
    except Exception as e:
        print(f"Error opening {input_path}: {e}", file=sys.stderr)
        return False

    sheet_names = wb.sheet_names()
    print(f"Processing: {input_path} ({len(sheet_names)} sheets)", file=sys.stderr)

    sheets_data = []
    for sheet_name in sheet_names:
        ws = wb.sheet_by_name(sheet_name)

        # Build merged cell map: (row, col) -> value or ''
        merged_values = {}
        for row_lo, row_hi, col_lo, col_hi in ws.merged_cells:
            top_left_value = ws.cell_value(row_lo, col_lo)
            for r in range(row_lo, row_hi):
                for c in range(col_lo, col_hi):
                    if r == row_lo and c == col_lo:
                        merged_values[(r, c)] = top_left_value
                    else:
                        merged_values[(r, c)] = ''

        all_rows = []
        for rx in range(ws.nrows):
            cells = []
            for cx in range(ws.ncols):
                if (rx, cx) in merged_values:
                    raw_val = merged_values[(rx, cx)]
                else:
                    raw_val = ws.cell_value(rx, cx)

                cell_type = ws.cell_type(rx, cx)

                # xlrd date handling: dates are stored as floats
                if cell_type == xlrd.XL_CELL_DATE:
                    try:
                        dt_tuple = xlrd.xldate_as_tuple(raw_val, wb.datemode)
                        dt = datetime(*dt_tuple)
                        cells.append(dt.strftime('%Y-%m-%d'))
                    except Exception:
                        cells.append(_format_cell_value(raw_val))
                elif cell_type == xlrd.XL_CELL_BOOLEAN:
                    cells.append('TRUE' if raw_val else 'FALSE')
                elif cell_type == xlrd.XL_CELL_EMPTY:
                    cells.append('')
                else:
                    cells.append(_format_cell_value(raw_val))
            all_rows.append(cells)

        sheets_data.append((sheet_name, all_rows))

    wb.release_resources()
    return _sheets_to_markdown(input_path, output_dir, config, args, sheets_data)
